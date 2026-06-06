"""Load Excel data and build conflict table."""
import openpyxl


def load_data(filepath: str) -> tuple[list, dict, dict]:
    """Read Excel file and return (courses list, conflict table, student courses dict)."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Read all course codes from the Course_Catalog sheet
    ws_catalog = wb["Course_Catalog"]
    courses = []
    for row in ws_catalog.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if code is not None:
            courses.append(str(code).strip())

    # Read which courses each student is enrolled in
    ws_enroll = wb["Enrollment_Pairs"]
    student_courses: dict[str, list[str]] = {}
    for row in ws_enroll.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        sid = str(row[0]).strip()
        course = str(row[1]).strip()
        if sid not in student_courses:
            student_courses[sid] = []
        student_courses[sid].append(course)

    wb.close()

    # Build conflict table: conflict_table[A][B] = number of students in both A and B
    conflict_table: dict[str, dict[str, int]] = {
        c: {d: 0 for d in courses} for c in courses
    }
    for sid, enrolled in student_courses.items():
        for i in range(len(enrolled)):
            for j in range(i + 1, len(enrolled)):
                a, b = enrolled[i], enrolled[j]
                if a in conflict_table and b in conflict_table.get(a, {}):
                    conflict_table[a][b] += 1  # Count students in both courses
                    conflict_table[b][a] += 1

    return courses, conflict_table, student_courses


def load_slots() -> list[str]:
    """Return all 18 exam slots (6 days × 3 slots per day)."""
    # Slots are named D1S1 to D6S3 where D=day and S=slot number
    return [f"D{d}S{s}" for d in range(1, 7) for s in range(1, 4)]


def load_bad_schedule(filepath: str) -> dict:
    """Read the bad example schedule from Excel and return course to slot mapping."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Sample_Bad_Schedule"]
    schedule = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        # Map each course code to its slot ID
        schedule[str(row[0]).strip()] = str(row[1]).strip()
    wb.close()
    return schedule
